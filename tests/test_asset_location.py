from __future__ import annotations

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from main import app
from modules.portfolio.services.advisory.tax_rules import public_registry, rules_as_of
from modules.portfolio.services.after_tax import estimate_after_tax
from modules.portfolio.services.asset_location import (
    check_account_eligibility,
    optimize_asset_location,
)
from modules.portfolio.services.tax_harvesting import evaluate_harvest
from modules.portfolio.services.tax_location_export import build_ca_workbook
from shared.config import APP_ROOT_PATH


AS_OF = "2026-08-29"


def _account(account_id: str, account_type: str, *, owner="owner-a", **profile):
    return {
        "account_id": account_id,
        "account_profile": {
            "owner_ref": owner,
            "account_type": account_type,
            "india_residency_status": profile.pop("residency", "RESIDENT"),
            "repatriability": profile.pop("repatriability", "REPATRIABLE"),
            "tax_lots_available": profile.pop("tax_lots_available", True),
            "estate_tax_review_status": profile.pop("estate_tax_review_status", "REVIEWED"),
            **profile,
        },
    }


def _evidence(capital_rate: float, withholding_rate: float = 0, **extra):
    return {
        "capital_gains_rate_pct": capital_rate,
        "withholding_rate_pct": withholding_rate,
        "source_url": "https://tax-authority.example/rule",
        "effective_from": "2026-01-01",
        "last_reviewed": "2026-08-01",
        "treaty_verified": True,
        **extra,
    }


def _candidate(**extra):
    return {
        "instrument_id": "fund-world",
        "symbol": "WORLD",
        "instrument_type": "etf",
        "fund_domicile": "IE",
        "security_country": "IE",
        "exact_product_id": "IE00TEST",
        "share_class": "ACC",
        "pre_tax_return_pct": {"bear": -10, "base": 10, "bull": 20},
        "dividend_yield_pct": 2,
        "tax_evidence": _evidence(10, 10),
        **extra,
    }


def test_same_etf_has_different_after_tax_outcome_by_account():
    candidate = _candidate(
        tax_evidence_by_account_type={
            "RESIDENT_DEMAT": _evidence(20, 10),
            "GLOBAL_BROKER": _evidence(5, 5),
        }
    )
    resident = estimate_after_tax(candidate, _account("india", "RESIDENT_DEMAT"), as_of=AS_OF)
    global_result = estimate_after_tax(
        candidate,
        _account("global", "GLOBAL_BROKER", residency="NON_RESIDENT"),
        as_of=AS_OF,
    )
    assert resident["status"] == global_result["status"] == "AVAILABLE"
    assert global_result["scenarios"]["base"]["after_tax_return_pct"] > resident["scenarios"]["base"]["after_tax_return_pct"]


def test_unverified_gift_product_requires_tax_review():
    result = estimate_after_tax(
        _candidate(),
        _account("gift", "GIFT_IBU", gift_product_tax_verified=False),
        as_of=AS_OF,
    )
    assert result["status"] == "TAX_REVIEW_REQUIRED"
    assert result["scenarios"]["base"] is None


def test_nri_indian_equity_sale_remains_indian_tax_relevant_and_tds_is_not_final():
    candidate = _candidate(instrument_type="equity", fund_domicile="", security_country="IN")
    result = estimate_after_tax(
        candidate,
        _account("nro", "NRO_NON_PIS", residency="NRI"),
        as_of=AS_OF,
    )
    assert "INDIA_TAX_RELEVANT" in result["flags"]
    assert "TDS_NOT_FINAL_LIABILITY" in result["flags"]


def test_resident_loss_planning_requires_fifo_lots():
    result = evaluate_harvest(
        {"symbol": "LOSS", "independent_sell_case": True, "last_price": 80},
        _account("india", "RESIDENT_DEMAT", tax_lots_available=False),
        lots=[],
        as_of=AS_OF,
    )
    assert result["status"] == "TAX_REVIEW_REQUIRED"
    assert "FIFO" in result["reason"]


def test_internal_family_move_is_not_assumed_tax_free():
    candidate = _candidate(
        current_account_id="one",
        tax_evidence_by_account_type={
            "RESIDENT_DEMAT": _evidence(25),
            "GLOBAL_BROKER": _evidence(5),
        },
    )
    result = optimize_asset_location(
        candidate,
        [
            _account("one", "RESIDENT_DEMAT", owner="owner-a"),
            _account("two", "GLOBAL_BROKER", owner="owner-b", residency="NON_RESIDENT"),
        ],
        as_of=AS_OF,
    )
    assert result["recommended_action"] == "USE_NEW_CONTRIBUTIONS_ELSEWHERE"
    assert result["transfer_assumed"] is False
    assert "Ownership differs" in result["reason"]


def test_broker_tds_is_never_treated_as_final_liability():
    result = evaluate_harvest(
        {"symbol": "INDIA", "independent_sell_case": True},
        _account("nro", "NRO_NON_PIS", residency="NRI"),
        lots=[],
        as_of=AS_OF,
    )
    assert result["tds_is_final_liability"] is False
    assert "not final liability" in " ".join(result["uncertainties"])


def test_us_situs_exposure_triggers_estate_review():
    result = estimate_after_tax(
        _candidate(fund_domicile="US", security_country="US"),
        _account("global", "GLOBAL_BROKER", residency="NON_RESIDENT", estate_tax_review_status="UNKNOWN"),
        as_of=AS_OF,
    )
    assert result["status"] == "TAX_REVIEW_REQUIRED"
    assert "US_SITUS_ESTATE_REVIEW" in result["flags"]


def test_exit_load_can_outweigh_location_benefit():
    candidate = _candidate(
        current_account_id="india",
        exit_load_pct=3,
        tax_evidence_by_account_type={
            "RESIDENT_DEMAT": _evidence(20),
            "GLOBAL_BROKER": _evidence(10),
        },
    )
    result = optimize_asset_location(
        candidate,
        [
            _account("india", "RESIDENT_DEMAT"),
            _account("global", "GLOBAL_BROKER", residency="NON_RESIDENT"),
        ],
        as_of=AS_OF,
    )
    assert result["recommended_action"] == "DO_NOT_MOVE_COST_EXCEEDS_BENEFIT"


def test_optimizer_respects_account_eligibility():
    account = _account("pis", "NRE_PIS", permitted_instrument_types=["equity"])
    eligibility = check_account_eligibility(account, _candidate(instrument_type="mutual_fund"))
    assert eligibility["eligible"] is False
    result = optimize_asset_location(_candidate(instrument_type="mutual_fund"), [account], as_of=AS_OF)
    assert result["selected_account_id"] is None


def test_unknown_treaty_evidence_blocks_certainty():
    candidate = _candidate(tax_evidence=_evidence(10, treaty_verified=False))
    result = estimate_after_tax(
        candidate,
        _account("global", "GLOBAL_BROKER", residency="NON_RESIDENT"),
        as_of=AS_OF,
    )
    assert result["status"] == "TAX_REVIEW_REQUIRED"


def test_rule_effective_dates_are_enforced():
    assert rules_as_of("2026-08-27") == []
    assert len(rules_as_of(AS_OF)) == len(public_registry(AS_OF)["rules"])


def test_ca_export_contains_sources_assumptions_lots_and_actions():
    workbook = build_ca_workbook(
        rules=public_registry(AS_OF)["rules"],
        assumptions=[{"name": "as_of", "value": AS_OF}],
        lots=[{"lot_id": "lot-1", "cost_basis": 100}],
        actions=[{"action": "REVIEW_WITH_CA"}],
    )
    parsed = load_workbook(BytesIO(workbook.read()), read_only=True)
    assert {"Read me", "Rules and sources", "Assumptions", "FIFO lots", "Proposed actions"} <= set(parsed.sheetnames)
    headers = [cell.value for cell in next(parsed["Rules and sources"].iter_rows())]
    assert "source_url" in headers


def test_api_v1_remains_compatible_and_tax_api_is_additive():
    client = TestClient(app)
    version = client.get(f"{APP_ROOT_PATH}/api/portfolio/version")
    rules = client.get(f"{APP_ROOT_PATH}/api/portfolio/tax/rules", params={"as_of": AS_OF})
    assert version.status_code == 200
    assert version.json()["contract_version"] == "2026-05-mobile-mvp-v1"
    assert rules.status_code == 200
    assert rules.json()["rules"]
