"""Sort, group, and export portfolio holdings for the UI."""

from __future__ import annotations

import json
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from modules.portfolio.config import get_account_code
from modules.portfolio.services.market_data import _METALS_SYMBOLS, is_us_exchange, normalize_symbol

_INDIAN_EXCHANGES = frozenset({"NSE", "BSE"})

SORT_FIELDS: dict[str, str] = {
    "value": "current_value",
    "pnl": "pnl",
    "pnl_pct": "pnl_pct",
    "symbol": "symbol",
    "sector": "sector",
    "cap": "market_cap",
    "pe": "pe_ratio",
    "pct_52w": "pct_from_52w_high",
    "upside": "upside_pct",
    "signal": "rating_rank",
    "qty": "quantity",
    "ltp": "last_price",
    "avg": "avg_price",
    "weight": "pct_of_portfolio",
    "decision": "decision_rank",
}

HOLDINGS_PAGE_SIZE = 50

CAP_GROUP_ORDER = ("Large", "Mid", "Small", "Multi-cap", "ETF", "Unclassified")
CAP_RANK = {label: index for index, label in enumerate(CAP_GROUP_ORDER)}

SIGNAL_GROUP_ORDER = (
    "External: strong buy",
    "External: buy",
    "External: hold",
    "External: sell",
    "External: strong sell",
    "External view unavailable",
)
SIGNAL_GROUP_RANK = {label: index for index, label in enumerate(SIGNAL_GROUP_ORDER)}

DECISION_GROUP_ORDER = (
    "Fix data first",
    "Not currently executable",
    "Tax review first",
    "Research before exiting",
    "Review for a possible trim",
    "Research before adding",
    "Exit",
    "Trim gradually",
    "Add more",
    "Add gradually",
    "Hold — no new money",
    "Hold — position full",
    "Hold",
    "Wait for evidence",
    "Decision unavailable",
)
DECISION_GROUP_RANK = {label: index for index, label in enumerate(DECISION_GROUP_ORDER)}

ASSET_CLASS_GROUP_ORDER = ("Equity", "US stocks", "Metals", "Crypto", "Mutual funds")
ASSET_CLASS_GROUP_RANK = {label: index for index, label in enumerate(ASSET_CLASS_GROUP_ORDER)}

_ASSET_CLASS_FILTER_KEYS: dict[str, str] = {
    "Equity": "equity",
    "US stocks": "us_stocks",
    "Metals": "metals",
    "Crypto": "crypto",
    "Mutual funds": "mf",
}

def account_abbrev(account_id: str | None) -> str:
    """Family dashboard account code (AB, RB, SB, HB)."""
    if not account_id:
        return "?"
    try:
        return get_account_code(account_id)
    except KeyError:
        return str(account_id)[:2].upper()


def _family_aggregate_key(holding: dict[str, Any]) -> str:
    """Bucket by canonical identity; never merge a symbol across exchanges heuristically."""
    symbol = normalize_symbol(holding.get("symbol") or "")
    if not symbol:
        return ""
    if holding.get("instrument_id"):
        return f"ID:{holding['instrument_id']}"
    if holding.get("isin"):
        return f"ISIN:{str(holding['isin']).upper()}"
    return f"SYMBOL:{(holding.get('exchange') or 'NSE').upper()}:{symbol}"


def _holding_row_key(holding: dict[str, Any], *, aggregated: bool) -> str:
    symbol = normalize_symbol(holding.get("symbol") or "?")
    if aggregated:
        return _family_aggregate_key(holding) or symbol
    exchange = holding.get("exchange") or "NSE"
    account_id = holding.get("account_id") or "single"
    return f"{symbol}:{exchange}:{account_id}"


def _account_breakdown_row(part: dict[str, Any], *, symbol: str | None) -> dict[str, Any]:
    qty = float(part.get("quantity") or 0)
    invested = float(part.get("invested") or 0)
    current = float(part.get("display_value") or part.get("marked_value") or part.get("current_value") or 0)
    pnl = float(
        part.get("display_pnl")
        if part.get("display_pnl") is not None
        else current - invested
    )
    return {
        "account_id": part.get("account_id"),
        "abbrev": account_abbrev(part.get("account_id")),
        "broker": part.get("broker"),
        "exchange": part.get("exchange"),
        "symbol": symbol,
        "quantity": qty,
        "avg_price": round((invested / qty) if qty else float(part.get("avg_price") or 0), 2),
        "invested": round(invested, 2),
        "current_value": round(current, 2),
        "pnl": round(pnl, 2),
        "pnl_pct": round((pnl / invested * 100) if invested else 0.0, 2),
    }


def _format_account_column(parts: list[dict[str, Any]]) -> str:
    """e.g. AB (100) + HB (300) or AB when a single account holds the symbol."""
    ordered = sorted(parts, key=lambda p: account_abbrev(p.get("account_id")))
    if len(ordered) == 1:
        return account_abbrev(ordered[0].get("account_id"))
    return " + ".join(
        f"{account_abbrev(p.get('account_id'))} ({int(p.get('quantity') or 0)})"
        for p in ordered
    )


def _merge_holding_group(parts: list[dict[str, Any]]) -> dict[str, Any]:
    """Combine the same symbol across accounts into one family row."""
    primary = max(
        parts,
        key=lambda p: float(
            p.get("display_value") or p.get("marked_value") or p.get("current_value") or 0
        ),
    )
    total_qty = sum(float(p.get("quantity") or 0) for p in parts)
    total_invested = sum(float(p.get("invested") or 0) for p in parts)
    total_current = sum(
        float(p.get("display_value") or p.get("marked_value") or p.get("current_value") or 0)
        for p in parts
    )
    total_pnl = total_current - total_invested
    avg_price = (total_invested / total_qty) if total_qty else 0.0
    broker_value = sum(float(p.get("broker_reported_value") or p.get("current_value") or 0) for p in parts)
    marked_value = total_current
    reconciliation_delta = sum(float(p.get("reconciliation_delta") or 0) for p in parts)
    state_rank = {
        "UNRESOLVED_IDENTITY": 5,
        "CORPORATE_ACTION_REVIEW": 4,
        "BLOCKING_MISMATCH": 3,
        "WARNING": 2,
        "RECONCILED_WITH_TIMING_DIFFERENCE": 1,
        "RECONCILED": 0,
    }
    states = [str(p.get("reconciliation_state")) for p in parts if p.get("reconciliation_state")]

    exchanges = sorted({(p.get("exchange") or "NSE").upper() for p in parts})
    if len(exchanges) == 1:
        exchange_label = exchanges[0]
    elif set(exchanges) <= _INDIAN_EXCHANGES:
        exchange_label = "NSE · BSE"
    else:
        exchange_label = " · ".join(exchanges)

    merged = dict(primary)
    display_price = (total_current / total_qty) if total_qty else 0.0
    merged.update(
        {
            "quantity": total_qty,
            "avg_price": round(avg_price, 2),
            "last_price": round(display_price, 4),
            "display_price": round(display_price, 4),
            "invested": round(total_invested, 2),
            "current_value": round(total_current, 2),
            "pnl": round(total_pnl, 2),
            "pnl_pct": round((total_pnl / total_invested * 100) if total_invested else 0.0, 2),
            "exchange": exchange_label,
            "account_label": _format_account_column(parts),
            "account_breakdown": [
                _account_breakdown_row(p, symbol=primary.get("symbol"))
                for p in sorted(parts, key=lambda p: account_abbrev(p.get("account_id")))
            ],
            "is_aggregated": len(parts) > 1,
            "account_id": None,
            "account_codes": ",".join(
                sorted({account_abbrev(p.get("account_id")) for p in parts})
            ),
            "broker_reported_price": round((broker_value / total_qty) if total_qty else 0.0, 4),
            "broker_reported_value": round(broker_value, 2),
            "market_price": round((marked_value / total_qty) if total_qty else 0.0, 4),
            "market_price_source": primary.get("market_price_source") or "broker_snapshot",
            "market_price_as_of": primary.get("market_price_as_of"),
            "marked_value": round(marked_value, 2),
            "reconciliation_delta": round(reconciliation_delta, 2),
            "reconciliation_state": (
                max(states, key=lambda value: state_rank.get(value, 2)) if states else None
            ),
            "reconciliation_blocking": any(
                bool(p.get("reconciliation_blocking")) for p in parts
            ),
        }
    )
    high_52w = merged.get("high_52w")
    target_price = merged.get("target_price")
    if display_price > 0 and high_52w:
        merged["pct_from_52w_high"] = round(
            ((display_price - float(high_52w)) / float(high_52w)) * 100,
            2,
        )
    if display_price > 0 and target_price:
        merged["upside_pct"] = round(
            ((float(target_price) - display_price) / display_price) * 100,
            2,
        )
    merged["holding_key"] = _holding_row_key(merged, aggregated=True)
    return merged


def holdings_financials_map(holdings: list[dict[str, Any]]) -> dict[str, dict[str, dict[str, float]]]:
    """Per holding_key → account code → invested/current_value for client-side filters."""
    out: dict[str, dict[str, dict[str, float]]] = {}
    for holding in holdings:
        key = holding.get("holding_key")
        if not key:
            continue
        breakdown = holding.get("account_breakdown")
        if breakdown:
            parts: dict[str, dict[str, float]] = {}
            for part in breakdown:
                abbrev = (part.get("abbrev") or "").strip()
                if not abbrev:
                    continue
                parts[abbrev] = {
                    "invested": round(float(part.get("invested") or 0), 2),
                    "current_value": round(float(part.get("current_value") or 0), 2),
                }
            if parts:
                out[key] = parts
            continue
        code = (
            holding.get("account_code")
            or holding.get("account_label")
            or (holding.get("account_codes") or "").split(",")[0].strip()
        )
        if code:
            out[key] = {
                str(code): {
                    "invested": round(float(holding.get("invested") or 0), 2),
                    "current_value": round(float(holding.get("current_value") or 0), 2),
                }
            }
    return out


def all_holdings_from_view(holdings_view: dict[str, Any]) -> list[dict[str, Any]]:
    """Flatten grouped or flat holdings_view for export/financials maps."""
    if holdings_view.get("mode") == "grouped" and holdings_view.get("groups"):
        rows: list[dict[str, Any]] = []
        for group in holdings_view["groups"]:
            rows.extend(group.get("holdings") or [])
        return rows
    return list(holdings_view.get("holdings") or [])


def filter_holdings_by_account_codes(
    holdings: list[dict[str, Any]],
    account_codes: list[str] | None,
) -> list[dict[str, Any]]:
    """Keep holdings whose account code is in account_codes. Empty/None = no filter."""
    if not account_codes:
        return holdings
    allowed = {code.strip().upper() for code in account_codes if code and str(code).strip()}
    if not allowed:
        return holdings
    filtered: list[dict[str, Any]] = []
    for holding in holdings:
        account_id = holding.get("account_id")
        code = (
            get_account_code(account_id)
            if account_id
            else (holding.get("account_code") or holding.get("account_label") or "")
        )
        if str(code).strip().upper() in allowed:
            filtered.append(holding)
    return filtered


def _attach_account_codes(holdings: list[dict[str, Any]]) -> None:
    """Set comma-separated account codes (AB, HB, …) for client-side filtering."""
    for holding in holdings:
        if holding.get("account_codes"):
            continue
        breakdown = holding.get("account_breakdown")
        if breakdown:
            holding["account_codes"] = ",".join(
                sorted({b.get("abbrev") for b in breakdown if b.get("abbrev")})
            )
        else:
            code = holding.get("account_code") or holding.get("account_label")
            if code:
                holding["account_codes"] = str(code)


def aggregate_holdings_across_accounts(holdings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """One row per symbol (NSE+BSE merged); Account shows AB (n) + HB (m)."""
    buckets: dict[str, list[dict[str, Any]]] = {}
    for holding in holdings:
        key = _family_aggregate_key(holding)
        if not key:
            continue
        buckets.setdefault(key, []).append(holding)

    merged_rows: list[dict[str, Any]] = []
    for parts in buckets.values():
        merged_rows.append(_merge_holding_group(parts))
    return merged_rows


def _signal_group_label(holding: dict[str, Any]) -> str:
    """Neutral external-consensus buckets; never portfolio actions."""
    label = holding.get("rating_label")
    mapping = {
        "Strong buy": "External: strong buy",
        "Buy": "External: buy",
        "Hold": "External: hold",
        "Sell": "External: sell",
        "Strong sell": "External: strong sell",
    }
    return mapping.get(label, "External view unavailable")


def _decision_group_label(holding: dict[str, Any]) -> str:
    presentation = holding.get("decision_presentation") or {}
    return str(presentation.get("label") or "Decision unavailable")


def _asset_class_group_label(holding: dict[str, Any]) -> str:
    """Dashboard asset bucket for allocation chart (Group by Class)."""
    if holding.get("asset_class") == "mf":
        return "Mutual funds"

    sector = (holding.get("sector") or "").strip()
    symbol = (holding.get("symbol") or "").upper()

    if (
        holding.get("asset_class") == "crypto"
        or sector == "Crypto"
        or symbol in ("BTC", "BTC-USD")
    ):
        return "Crypto"
    if sector == "Metals" or normalize_symbol(symbol) in _METALS_SYMBOLS:
        return "Metals"
    if is_us_exchange(holding.get("exchange")) or holding.get("broker") == "sarwa":
        return "US stocks"
    return "Equity"


def _asset_class_filter_key(holding: dict[str, Any]) -> str:
    """Value for data-asset-class / Show filter checkboxes."""
    return _ASSET_CLASS_FILTER_KEYS.get(_asset_class_group_label(holding), "equity")


def _annotate_asset_class_groups(holdings: list[dict[str, Any]]) -> None:
    for holding in holdings:
        holding["asset_class_group"] = _asset_class_group_label(holding)
        holding["asset_class_filter"] = _asset_class_filter_key(holding)


def _cap_group_label(holding: dict[str, Any]) -> str:
    """Cap bucket for grouping; debt MFs and unknowns → Unclassified."""
    cap = holding.get("market_cap")
    if cap in (None, "", "MF"):
        if holding.get("asset_class") == "mf":
            return "Unclassified"
        return "ETF"
    return cap


def _group_label(field: str, value: str | None) -> str:
    if field == "sector":
        return value or "Unclassified"
    return value or "Unclassified"


def _annotate_portfolio_weights(holdings: list[dict[str, Any]]) -> None:
    """Set pct_of_portfolio = share of total current value (e.g. 1.8 for 1.8%)."""
    total_value = sum(float(h.get("current_value") or 0) for h in holdings)
    for holding in holdings:
        value = float(holding.get("current_value") or 0)
        holding["pct_of_portfolio"] = round(
            (value / total_value * 100) if total_value else 0.0,
            2,
        )


def _has_sort_value(holding: dict[str, Any], field: str) -> bool:
    """Whether a holding has a real value for sorting (skip — rows)."""
    if field == "market_cap":
        return True
    if field == "rating_rank":
        return holding.get("rating_label") is not None
    if field == "decision_rank":
        return bool((holding.get("decision_presentation") or {}).get("label"))
    return holding.get(field) is not None


def _sort_value(holding: dict[str, Any], field: str, *, order: str = "desc") -> Any:
    """Comparable sort value for holdings that have data."""
    reverse = order.lower() != "asc"

    if field == "market_cap":
        rank = CAP_RANK.get(holding.get("market_cap") or "ETF", len(CAP_RANK))
        return (rank, -(holding.get("current_value") or 0))

    if field == "rating_rank":
        rank = holding.get("rating_rank", 2)
        upside = holding.get("upside_pct")
        if reverse:
            # desc: Strong buy first, then higher upside within the same signal
            upside_key = upside if upside is not None else float("-inf")
        else:
            # asc: Strong sell first, then lower upside within the same signal
            upside_key = upside if upside is not None else float("inf")
        return (-rank, upside_key)

    if field == "decision_rank":
        label = _decision_group_label(holding)
        rank = DECISION_GROUP_RANK.get(label, len(DECISION_GROUP_RANK))
        return -rank if reverse else rank

    value = holding.get(field)
    if isinstance(value, str):
        return value.lower()
    return float(value)


def sort_holdings(
    holdings: list[dict[str, Any]],
    *,
    sort: str = "value",
    order: str = "desc",
) -> list[dict[str, Any]]:
    """Sort holdings. Default: current value high to low. Missing metrics sort last."""
    field = SORT_FIELDS.get(sort, "current_value")
    reverse = order.lower() != "asc"

    with_value = [h for h in holdings if _has_sort_value(h, field)]
    without_value = [h for h in holdings if not _has_sort_value(h, field)]

    with_value.sort(key=lambda h: _sort_value(h, field, order=order), reverse=reverse)
    return with_value + without_value


def _group_summary(holdings: list[dict[str, Any]]) -> dict[str, Any]:
    total_invested = sum(h.get("invested") or 0 for h in holdings)
    total_current_value = sum(h.get("current_value") or 0 for h in holdings)
    total_pnl = sum(h.get("pnl") or 0 for h in holdings)
    return {
        "holdings_count": len(holdings),
        "total_invested": round(total_invested, 2),
        "total_current_value": round(total_current_value, 2),
        "total_pnl": round(total_pnl, 2),
        "total_pnl_pct": round((total_pnl / total_invested * 100) if total_invested else 0.0, 2),
    }


def _expand_holdings_by_account(holdings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Split aggregated family rows into one row per account for group-by-account."""
    expanded: list[dict[str, Any]] = []
    for holding in holdings:
        breakdown = holding.get("account_breakdown")
        if not breakdown:
            expanded.append(holding)
            continue
        for part in breakdown:
            abbrev = part.get("abbrev") or "?"
            qty = float(part.get("quantity") or 0)
            invested = float(part.get("invested") or 0)
            current = float(part.get("current_value") or 0)
            pnl = current - invested
            row = dict(holding)
            row.update(
                {
                    "quantity": qty,
                    "invested": round(invested, 2),
                    "current_value": round(current, 2),
                    "pnl": round(pnl, 2),
                    "pnl_pct": round((pnl / invested * 100) if invested else 0.0, 2),
                    "account_label": abbrev,
                    "account_codes": abbrev,
                    "account_breakdown": None,
                    "is_aggregated": False,
                    "holding_key": f"{holding.get('holding_key', holding.get('symbol'))}:{abbrev}",
                }
            )
            expanded.append(row)
    return expanded


def group_holdings(
    holdings: list[dict[str, Any]],
    *,
    group_by: str,
    sort: str = "value",
    order: str = "desc",
) -> list[dict[str, Any]]:
    """Group holdings by decision, cap, sector, account, external view, or class."""
    if group_by == "decision":
        field = "decision_group"
    elif group_by == "cap":
        field = "market_cap"
    elif group_by == "sector":
        field = "sector"
    elif group_by == "account":
        field = "account_label"
    elif group_by == "signal":
        field = "signal_group"
    elif group_by == "asset_class":
        field = "asset_class_group"
    else:
        return []

    source = _expand_holdings_by_account(holdings) if group_by == "account" else holdings

    buckets: dict[str, list[dict[str, Any]]] = {}
    for holding in source:
        if field == "decision_group":
            label = _decision_group_label(holding)
        elif field == "market_cap":
            label = _cap_group_label(holding)
        elif field == "account_label":
            label = holding.get("account_label") or holding.get("account_code") or "?"
        elif field == "signal_group":
            label = _signal_group_label(holding)
        elif field == "asset_class_group":
            label = _asset_class_group_label(holding)
        else:
            label = _group_label(field, holding.get(field))
        buckets.setdefault(label, []).append(holding)

    def group_rank(label: str) -> tuple:
        if field == "decision_group":
            return (DECISION_GROUP_RANK.get(label, len(DECISION_GROUP_RANK)), label.lower())
        if field == "market_cap":
            return (CAP_RANK.get(label, len(CAP_RANK)), label.lower())
        if field == "account_label":
            order_codes = ["AB", "RB", "SB", "HB"]
            rank = order_codes.index(label) if label in order_codes else 99
            return (rank, label.lower())
        if field == "signal_group":
            return (SIGNAL_GROUP_RANK.get(label, len(SIGNAL_GROUP_RANK)), label.lower())
        if field == "asset_class_group":
            return (ASSET_CLASS_GROUP_RANK.get(label, len(ASSET_CLASS_GROUP_RANK)), label.lower())
        if label == "Unclassified":
            return (1, label.lower())
        return (0, label.lower())

    groups = []
    for label in sorted(buckets, key=group_rank):
        items = sort_holdings(buckets[label], sort=sort, order=order)
        groups.append(
            {
                "label": label,
                "holdings": items,
                "summary": _group_summary(items),
            }
        )
    if field == "decision_group":
        groups.sort(
            key=lambda group: DECISION_GROUP_RANK.get(group["label"], len(DECISION_GROUP_RANK)),
        )
    elif field == "signal_group":
        groups.sort(
            key=lambda group: SIGNAL_GROUP_RANK.get(group["label"], len(SIGNAL_GROUP_RANK)),
        )
    else:
        groups.sort(
            key=lambda group: group["summary"]["total_current_value"],
            reverse=True,
        )
    return groups


def prepare_holdings_view(
    holdings: list[dict[str, Any]],
    *,
    sort: str = "value",
    order: str = "desc",
    group_by: str | None = None,
    aggregate_across_accounts: bool = False,
) -> dict[str, Any]:
    """Return flat or grouped holdings ready for templates/export."""
    if aggregate_across_accounts:
        holdings = aggregate_holdings_across_accounts(holdings)
    else:
        for holding in holdings:
            holding["holding_key"] = _holding_row_key(holding, aggregated=False)
            if holding.get("account_id"):
                holding["account_label"] = account_abbrev(holding["account_id"])

    _annotate_portfolio_weights(holdings)
    _annotate_asset_class_groups(holdings)
    _attach_account_codes(holdings)
    sorted_holdings = sort_holdings(holdings, sort=sort, order=order)
    total_count = len(sorted_holdings)

    if group_by in ("decision", "cap", "sector", "account", "signal", "asset_class"):
        groups = group_holdings(sorted_holdings, group_by=group_by, sort=sort, order=order)
        portfolio_total_value = sum(h.get("current_value") or 0 for h in sorted_holdings)
        for group in groups:
            group_value = group["summary"]["total_current_value"]
            group["summary"]["pct_of_portfolio"] = round(
                (group_value / portfolio_total_value * 100) if portfolio_total_value else 0.0,
                1,
            )
        overview_labels = [g["label"] for g in groups]
        overview_values = [g["summary"]["total_current_value"] for g in groups]
        overview_pcts = [g["summary"]["pct_of_portfolio"] for g in groups]
        return {
            "mode": "grouped",
            "holdings": sorted_holdings,
            "groups": groups,
            "sort": sort,
            "order": order,
            "group_by": group_by,
            "total_count": total_count,
            "page_size": HOLDINGS_PAGE_SIZE,
            "overview_chart_json": json.dumps(
                {
                    "labels": overview_labels,
                    "values": overview_values,
                    "pcts": overview_pcts,
                }
            ),
        }

    return {
        "mode": "flat",
        "holdings": sorted_holdings,
        "groups": None,
        "sort": sort,
        "order": order,
        "group_by": None,
        "total_count": total_count,
        "page_size": HOLDINGS_PAGE_SIZE,
    }


EXPORT_COLUMN_HEADERS: dict[str, str] = {
    "instrument_id": "Instrument ID",
    "isin": "ISIN",
    "symbol": "Symbol",
    "exchange": "Exchange",
    "account": "Account",
    "cap": "Cap",
    "pe": "P/E",
    "sector": "Sector",
    "pct52w": "52W Δ %",
    "upside": "Upside %",
    "signal": "Signal",
    "weight": "% of total",
    "qty": "Qty",
    "avg": "Avg price",
    "ltp": "LTP",
    "value": "Value",
    "invested": "Invested",
    "pnl": "P&L",
    "pnl_pct": "P&L %",
    "broker_price": "Broker price",
    "broker_value": "Broker value",
    "market_price": "Market price",
    "marked_value": "Marked value",
    "price_source": "Market price source",
    "price_as_of": "Market price as of",
    "session_date": "Market session",
    "reconciliation_delta": "Reconciliation Δ",
    "reconciliation_state": "Data-quality state",
    "decision": "Portfolio decision",
    "readiness": "Decision readiness",
    "decision_confidence": "Decision confidence band",
    "decision_review_trigger": "Decision review trigger",
    "external_consensus": "External analyst consensus (context only)",
    "external_target_gap": "External target gap % (context only)",
}

EXPORT_COLUMN_ORDER: tuple[str, ...] = tuple(EXPORT_COLUMN_HEADERS.keys())


def _export_cell(col_id: str, holding: dict[str, Any]) -> Any:
    presentation = holding.get("decision_presentation") or {}
    external = holding.get("external_analyst_view") or {}
    if col_id == "instrument_id":
        return holding.get("instrument_id")
    if col_id == "isin":
        return holding.get("isin")
    if col_id == "symbol":
        return holding.get("symbol")
    if col_id == "exchange":
        return holding.get("exchange")
    if col_id == "account":
        return holding.get("account_label")
    if col_id == "cap":
        return holding.get("market_cap")
    if col_id == "pe":
        return holding.get("pe_ratio")
    if col_id == "sector":
        return holding.get("sector")
    if col_id == "pct52w":
        return holding.get("pct_from_52w_high")
    if col_id == "upside":
        return holding.get("upside_pct")
    if col_id == "signal":
        return holding.get("rating_label")
    if col_id == "weight":
        return holding.get("pct_of_portfolio")
    if col_id == "qty":
        return holding.get("quantity")
    if col_id == "avg":
        return holding.get("avg_price")
    if col_id == "ltp":
        return holding.get("last_price")
    if col_id == "value":
        return holding.get("current_value")
    if col_id == "invested":
        return holding.get("invested")
    if col_id == "pnl":
        return holding.get("pnl")
    if col_id == "pnl_pct":
        return holding.get("pnl_pct")
    if col_id == "broker_price":
        return holding.get("broker_reported_price")
    if col_id == "broker_value":
        return holding.get("broker_reported_value")
    if col_id == "market_price":
        return holding.get("market_price")
    if col_id == "marked_value":
        return holding.get("marked_value")
    if col_id == "price_source":
        return (holding.get("reconciliation") or {}).get("market_price_source")
    if col_id == "price_as_of":
        return (holding.get("reconciliation") or {}).get("market_price_as_of")
    if col_id == "session_date":
        return (holding.get("reconciliation") or {}).get("market_session_date")
    if col_id == "reconciliation_delta":
        return holding.get("reconciliation_delta")
    if col_id == "reconciliation_state":
        return holding.get("reconciliation_state")
    if col_id == "decision":
        return presentation.get("label")
    if col_id == "readiness":
        return presentation.get("readiness_label")
    if col_id == "decision_confidence":
        return presentation.get("confidence_band")
    if col_id == "decision_review_trigger":
        return presentation.get("review_trigger")
    if col_id == "external_consensus":
        return external.get("consensus_label")
    if col_id == "external_target_gap":
        return external.get("target_gap_pct")
    return None


def default_export_columns(*, include_account: bool) -> list[str]:
    if include_account:
        return list(EXPORT_COLUMN_ORDER)
    return [c for c in EXPORT_COLUMN_ORDER if c != "account"]


def normalize_export_columns(
    columns: list[str] | str | None,
    *,
    include_account: bool,
) -> list[str]:
    """Return ordered, valid export column ids (defaults to all allowed)."""
    allowed = set(default_export_columns(include_account=include_account))
    if not columns:
        return [c for c in EXPORT_COLUMN_ORDER if c in allowed]

    if isinstance(columns, str):
        requested = [part.strip() for part in columns.split(",") if part.strip()]
    else:
        requested = list(columns)

    picked = [c for c in EXPORT_COLUMN_ORDER if c in requested and c in allowed]
    return picked or [c for c in EXPORT_COLUMN_ORDER if c in allowed]


def export_column_options(*, include_account: bool) -> list[dict[str, str]]:
    return [
        {"id": col_id, "label": EXPORT_COLUMN_HEADERS[col_id]}
        for col_id in default_export_columns(include_account=include_account)
    ]


def _excel_row(holding: dict[str, Any], columns: list[str]) -> list[Any]:
    return [_export_cell(col_id, holding) for col_id in columns]


def build_holdings_excel(
    view: dict[str, Any],
    *,
    columns: list[str] | str | None = None,
    include_account: bool = False,
    sheet_title: str = "Holdings",
) -> bytes:
    """Build an .xlsx workbook from a prepared holdings view."""
    col_ids = normalize_export_columns(columns, include_account=include_account)
    headers = [EXPORT_COLUMN_HEADERS[c] for c in col_ids]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    if view["mode"] == "grouped" and view["groups"]:
        for group in view["groups"]:
            ws.append([f"{group['label']} ({group['summary']['holdings_count']})"])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            for holding in group["holdings"]:
                ws.append(_excel_row(holding, col_ids))
    else:
        for holding in view["holdings"]:
            ws.append(_excel_row(holding, col_ids))

    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 14
    ws.column_dimensions["A"].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
