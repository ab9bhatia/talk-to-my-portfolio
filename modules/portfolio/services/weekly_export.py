"""Excel export for weekly portfolio snapshots."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from modules.portfolio.db import weekly_history


def build_weekly_history_excel(
    *,
    scope: str = "family",
    account_id: str | None = None,
    weeks: int = 52,
) -> bytes:
    """Workbook: weekly totals, all position rows, and latest week-over-week changes."""
    snaps = weekly_history.list_snapshots(scope=scope, account_id=account_id, limit=weeks)
    if not snaps:
        raise ValueError("No weekly snapshots stored yet")

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Weekly summary"

    summary_headers = [
        "Week start",
        "Source",
        "Holdings",
        "Invested (INR)",
        "Current (INR)",
        "P&L (INR)",
        "P&L %",
        "USD/INR",
        "Notes",
    ]
    ws_summary.append(summary_headers)
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)

    for row in reversed(snaps):
        ws_summary.append(
            [
                row.get("week_start"),
                row.get("source"),
                row.get("holdings_count"),
                row.get("total_invested"),
                row.get("total_current"),
                row.get("total_pnl"),
                row.get("total_pnl_pct"),
                row.get("usd_inr"),
                row.get("notes"),
            ]
        )

    ws_positions = wb.create_sheet("Positions by week")
    pos_headers = [
        "Week start",
        "Symbol",
        "Exchange",
        "Asset class",
        "Qty",
        "Avg price",
        "LTP",
        "Invested",
        "Value",
        "P&L",
        "P&L %",
        "Sector",
        "Cap",
        "P/E",
        "Signal",
    ]
    ws_positions.append(pos_headers)
    for cell in ws_positions[1]:
        cell.font = Font(bold=True)

    for snap_row in reversed(snaps):
        detail = weekly_history.get_snapshot(snap_row["id"])
        if not detail:
            continue
        week = detail["week_start"]
        for p in detail.get("positions") or []:
            ws_positions.append(
                [
                    week,
                    p.get("symbol"),
                    p.get("exchange"),
                    p.get("asset_class"),
                    p.get("quantity"),
                    p.get("avg_price"),
                    p.get("last_price"),
                    p.get("invested"),
                    p.get("current_value"),
                    p.get("pnl"),
                    p.get("pnl_pct"),
                    p.get("sector"),
                    p.get("market_cap"),
                    p.get("pe_ratio"),
                    p.get("rating_label"),
                ]
            )

    compare = weekly_history.compare_weeks(scope=scope, account_id=account_id)
    ws_changes = wb.create_sheet("Week over week")
    change_headers = [
        "Change",
        "Symbol",
        "Exchange",
        "Qty before",
        "Qty after",
        "Qty delta",
        "Current week",
        "Previous week",
    ]
    ws_changes.append(change_headers)
    for cell in ws_changes[1]:
        cell.font = Font(bold=True)
    for ch in compare.get("changes") or []:
        ws_changes.append(
            [
                ch.get("change"),
                ch.get("symbol"),
                ch.get("exchange"),
                ch.get("qty_before"),
                ch.get("qty_after"),
                ch.get("qty_delta"),
                compare.get("week_start"),
                compare.get("previous_week"),
            ]
        )

    for ws in (ws_summary, ws_positions, ws_changes):
        for idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 14
        ws.column_dimensions["A"].width = 12
        if ws == ws_positions:
            ws.column_dimensions["B"].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
