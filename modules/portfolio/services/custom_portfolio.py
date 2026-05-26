"""Parse and serve custom (CSV / Excel) portfolio imports."""

from __future__ import annotations

import csv
import io
import re
from typing import Any

from modules.portfolio.config import get_custom_account, get_account_code
from modules.portfolio.db import custom_holdings as custom_db
from modules.portfolio.services.portfolio import summarize_holdings

_ALIASES = {
    "symbol": {"symbol", "ticker", "scrip", "stock", "name"},
    "quantity": {"quantity", "qty", "units", "shares", "holding"},
    "avg_price": {"avg_price", "average", "avg", "buy_price", "cost", "avg_cost"},
    "last_price": {"last_price", "ltp", "price", "current_price", "cmp", "market_price"},
    "exchange": {"exchange", "exch"},
}


def _norm_header(name: str) -> str | None:
    key = re.sub(r"[^a-z0-9]", "", (name or "").lower())
    for field, aliases in _ALIASES.items():
        if key in aliases:
            return field
    return None


def parse_tabular_rows(headers: list[str], data_rows: list[list[Any]]) -> list[dict[str, Any]]:
    """Map spreadsheet columns to holding rows."""
    mapping: dict[int, str] = {}
    for idx, header in enumerate(headers):
        field = _norm_header(str(header or ""))
        if field:
            mapping[idx] = field
    if "symbol" not in mapping.values():
        raise ValueError("Could not find a Symbol column (symbol, ticker, scrip, …)")

    rows: list[dict[str, Any]] = []
    for raw in data_rows:
        if not raw or all(str(c or "").strip() == "" for c in raw):
            continue
        row: dict[str, Any] = {}
        for idx, field in mapping.items():
            if idx < len(raw):
                row[field] = raw[idx]
        symbol = str(row.get("symbol") or "").strip().upper()
        if not symbol:
            continue
        try:
            qty = float(str(row.get("quantity") or "0").replace(",", ""))
        except ValueError:
            continue
        if qty <= 0:
            continue
        avg = row.get("avg_price")
        ltp = row.get("last_price")
        try:
            avg_f = float(str(avg).replace(",", "")) if avg not in (None, "") else None
        except ValueError:
            avg_f = None
        try:
            ltp_f = float(str(ltp).replace(",", "")) if ltp not in (None, "") else None
        except ValueError:
            ltp_f = None
        rows.append({
            "symbol": symbol,
            "exchange": str(row.get("exchange") or "NSE").upper().strip() or "NSE",
            "quantity": qty,
            "avg_price": avg_f or ltp_f or 0.0,
            "last_price": ltp_f or avg_f or 0.0,
        })
    if not rows:
        raise ValueError("No valid holdings found in file")
    return rows


def parse_csv_bytes(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    if not all_rows:
        raise ValueError("Empty CSV file")
    return parse_tabular_rows(all_rows[0], all_rows[1:])


def parse_xlsx_bytes(content: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty spreadsheet")
    headers = [str(c) if c is not None else "" for c in rows[0]]
    return parse_tabular_rows(headers, [list(r) for r in rows[1:]])


def import_file(
    account_id: str,
    content: bytes,
    *,
    filename: str,
    notes: str | None = None,
) -> dict[str, Any]:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        holdings = parse_csv_bytes(content)
        source = "csv"
    elif lower.endswith((".xlsx", ".xls")):
        holdings = parse_xlsx_bytes(content)
        source = "excel"
    elif lower.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif")):
        from modules.portfolio.services.holdings_screenshot import parse_holdings_screenshot

        media = "image/png"
        if lower.endswith(".jpg") or lower.endswith(".jpeg"):
            media = "image/jpeg"
        elif lower.endswith(".webp"):
            media = "image/webp"
        parsed = parse_holdings_screenshot(content, media_type=media)
        holdings = parsed["rows"]
        source = "screenshot"
        notes = notes or parsed.get("notes")
    else:
        raise ValueError("Supported: .csv, .xlsx, or portfolio screenshot (.png, .jpg)")

    count = custom_db.replace_holdings(account_id, holdings, source=source, notes=notes)
    return {"account_id": account_id, "imported": count, "source": source}


def fetch_custom_portfolio_live(account_id: str, *, with_metrics: bool = True) -> dict:
    """Custom account holdings from SQLite."""
    from modules.portfolio.services.market_data import apply_holdings_metric_overrides, enrich_holdings

    account = get_custom_account(account_id)
    holdings = custom_db.list_holdings(account_id)
    if with_metrics and holdings:
        holdings = enrich_holdings(holdings)
        apply_holdings_metric_overrides(holdings)
    return {
        "account_id": account_id,
        "account_code": account["code"],
        "account_label": account.get("label") or account["code"],
        "user_id": "custom",
        "broker": "custom",
        "summary": summarize_holdings(holdings),
        "holdings": holdings,
    }
