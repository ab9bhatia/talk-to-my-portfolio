"""CA-review workbook for asset-location planning; never an ITR."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font


def build_ca_workbook(
    *, rules: list[dict[str, Any]], assumptions: list[dict[str, Any]],
    lots: list[dict[str, Any]], actions: list[dict[str, Any]],
) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "Rules and sources", rules)
    _sheet(workbook, "Assumptions", assumptions)
    _sheet(workbook, "FIFO lots", lots)
    _sheet(workbook, "Proposed actions", actions)
    notice = workbook.create_sheet("Read me", 0)
    notice.append(["Purpose", "Planning package for CA review; not an ITR or filing calculation."])
    notice.append(["Execution", "Disabled. No transfer or broker order is created."])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(title)
    columns = sorted({key for row in rows for key in row}) if rows else ["No data"]
    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([
            str(row.get(column)) if isinstance(row.get(column), (dict, list, tuple)) else row.get(column)
            for column in columns
        ])
    sheet.freeze_panes = "A2"
